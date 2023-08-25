import os
import math
from datetime import datetime
from openpyxl import load_workbook
from Models import base
from sqlalchemy import func
from Util.Paths import pathFromCloudAS,appSheetFolderName,excelName,pedidoExcelSheet,UEN,imagesFolderName
from Util.Paths import version,TIPO_AUDITORIA,UEN,PROCESO,SUBPROCESO,vsMateriales,SUBPROCESOMateriales ,SUBPROCESOParametros,vsParametros
from Util.Paths import pathOneDrive,appSheetFolderName,pedidoExcelSheet
from Util.Paths import PkgName, AU_Identifier
from sqlalchemy import and_, case, exists, not_, or_, select
from Models.base import Base,session
import pandas as pd
from Util.Functions import datetime2ole
from Modules import log


def truncateTmpTable(strTmpTable:str):
    try:
        sql = 'TRUNCATE TABLE {0}'.format(strTmpTable)
    ###### Inicia la session de manera que todo lo que este identado dentro de este with haga rollback automaticamente si 
    ###### algo falla
        with session.begin():
            session.execute(sql) 
            session.commit()

        log.insertLogToDB('Success','truncateTmpTable','Success execution in Mod Audit_Units')
    except Exception as e:
        log.insertLogToDB(str(e),'truncateTmpTable','Failed execution in Mod Audit_Units')
        raise
        


def readExcelSrc(date:datetime):
    try:
        #####Obtiene el path donde esta el excel y lo crea dinamicamente
        path=os.path.sep.join([pathFromCloudAS, appSheetFolderName+"_{0}_{1}_{2}".format(date.day,date.month,date.year),excelName])

        ##### Carga el excel y genera un dataframe
        #df=pd.read_excel(path,sheet_name=pedidoExcelSheet,converters={'HORA_DE_INICIO':str,'HORA_DE_FINALIZACION':str})
        df=pd.read_excel(path,sheet_name=pedidoExcelSheet)
        
        return df
        

    except Exception as e:
        log.insertLogToDB(str(e),'readExcelTmp','Failed execution in Mod Audit_Units')
        raise


def insertUAToTmpTable(AuditData:pd.DataFrame,date:datetime,DimCampos,TmpTable):
    try:

        listAudits=[]
        fecha_carga=datetime.today()
        ###### Inicia la session de manera que todo lo que este identado dentro de este with haga rollback automaticamente si 
        ###### algo falla
        with session.begin():
            campos=findAllByTIPO_CAMPO(('GENERAL','PRUEBA','INTERNO'),DimCampos)
            ###### recibe los datos cargados desde el excel y recorre el dataframe para extraer y crear los objetos a insertar en la base
            ###### de datos. luego los agrega en la lista a insertar 
            for index, row in AuditData.iterrows():
                _AuditData=TmpTable()
                for campo in campos:
                    if(campo.NOMBRE_EN_BD not in ['ESTADO_GENERAL','VAL_HORA_FINAL','TIEMPO_EN_MIN','FECHA_CARGA','AÑO','MES','FECHA'] and campo.TIPO_DATO not in ['HORA','IMAGEN']):
                        setattr(_AuditData,campo.NOMBRE_EN_BD,  None if pd.isnull(getattr(row,campo.NOMBRE_EN_EXCEL.strip())) else getattr(row,campo.NOMBRE_EN_EXCEL.strip()))

                    
                    elif(campo.TIPO_DATO=='HORA'):
                        setattr(_AuditData,campo.NOMBRE_EN_BD,  None if pd.isnull(getattr(row,campo.NOMBRE_EN_EXCEL.strip().strip())) else math.modf(datetime2ole(getattr(row,campo.NOMBRE_EN_EXCEL.strip())))[0])

                    elif(campo.NOMBRE_EN_BD=='VAL_HORA_FINAL' ):
                        setattr(_AuditData,campo.NOMBRE_EN_BD,  None if pd.isnull(getattr(row,campo.NOMBRE_EN_EXCEL.strip())) else datetime2ole(getattr(row,campo.NOMBRE_EN_EXCEL.strip())))
                    elif(campo.TIPO_DATO=='IMAGEN' ):
                        setattr(_AuditData,campo.NOMBRE_EN_BD,  None if pd.isnull(getattr(row,campo.NOMBRE_EN_EXCEL.strip())) else getattr(row,campo.NOMBRE_EN_EXCEL.strip()).replace(imagesFolderName,''))
                    
                '''
                listAudits.ESTADO_GENERAL= 'SIN GESTION' if listAudits.ESTADO_AUDITORIA==None and listAudits.TIMESTAMP==None else ('AUDITADO' if listAudits.ESTADO_AUDITORIA =='AUDITABLE' else 'NO AUDITADO')
                listAudits.TIEMPO_EN_MIN =None if listAudits.HORA_FINALIZACION== None or listAudits.HORA_INICIO== None else (listAudits.HORA_FINALIZACION-listAudits.HORA_INICIO)*24*60      
                listAudits.FECHA_CARGA=fecha_carga    
                listAudits.AÑO=date.year
                listAudits.MES=date.month
                listAudits.FECHA=date   
                listAudits.MESA_ASIG=listAudits.MESA
                '''
                
                _AuditData=populationFieldsOfBusiness(_AuditData,date,fecha_carga)
                listAudits.append(_AuditData)
                #print(index,_AuditData.PEDIDO_ID,_AuditData.ESTADO_GENERAL)
            
            
        
            ######  Agrego la lista a la session para guardarlos en la base de datos y con el commit hago persistencia en la base de datos
            session.bulk_save_objects(listAudits)
            session.commit()
            log.insertLogToDB('Success','insertUAToTmpTable','Success execution in Mod Audit_Units')
    except Exception as e:
        log.insertLogToDB(str(e),'insertUAToTmpTable','Failed execution in Mod Audit_Units')
        raise



def findAllByTIPO_CAMPO(TIPO_CAMPO:tuple,DimCampos):
    try:
   
        params=session.query(DimCampos).filter(DimCampos.VS==version ,
                                                DimCampos.TIPO_AUDITORIA== TIPO_AUDITORIA,
                                                DimCampos.UEN== UEN,
                                                DimCampos.PROCESO== PROCESO,
                                                DimCampos.SUBPROCESO== SUBPROCESO,
                                                DimCampos.TIPO_CAMPO.in_(TIPO_CAMPO)).all()
            
        return params

        
    except Exception as e:
        log.insertLogToDB(str(e),'findAllByTIPO_CAMPO','Failed execution in Mod Audit_Units')
        raise

def prepareFieldsToLoad(df:pd.DataFrame):

    try:
    ##### renombra las columnas que tienen espacios para que sea mas facil trabajar con ellas
        headers=[]
        index=df.columns.array
        [headers.append(column.strip().replace(" ","_")) for column in index]
        
        df.columns=headers
        
        
        ##### GESTION DE FECHAS

        ####### FECHA_CARGA
        if(df.FECHA_CARGA.dtype == 'int64'):
            df.FECHA_CARGA = pd.to_datetime(df.FECHA_CARGA, unit='D', origin='1899-12-30T00:00:00')
        else:
            df.FECHA_CARGA = pd.to_datetime(df.FECHA_CARGA)
        ###### FECHA
        if(df.FECHA.dtype == 'int64'):
            df.FECHA = pd.to_datetime(df.FECHA, unit='D', origin='1899-12-30T00:00:00')
        else:
            df.FECHA = pd.to_datetime(df.FECHA)

        ###### FECHA_AUDITORIA
        if(df.FECHA_DE_AUDITORIA.dtype == 'int64'):
            df.FECHA_DE_AUDITORIA = pd.to_datetime(df.FECHA_DE_AUDITORIA, unit='D', origin='1899-12-30T00:00:00')
        else:
            df.FECHA_DE_AUDITORIA = pd.to_datetime(df.FECHA_DE_AUDITORIA)

        ###### TIMESTAMP
        if(df.TIMESTAMP.dtype == 'int64'):
            df.TIMESTAMP = pd.to_datetime(df.TIMESTAMP, unit='D', origin='1899-12-30T00:00:00')
        else:
            df.TIMESTAMP = pd.to_datetime(df.TIMESTAMP)
        
        ####### VAL_HORA_FINALIZACION

        if(df.VAL_HORA_FINALIZACION.dtype == 'int64'):
            df.VAL_HORA_FINALIZACION = pd.to_datetime(df.VAL_HORA_FINALIZACION, unit='D', origin='1899-12-30T00:00:00')
        else:
            df.VAL_HORA_FINALIZACION = pd.to_datetime(df.VAL_HORA_FINALIZACION)

        df.HORA_DE_INICIO=pd.to_datetime(df.HORA_DE_INICIO, format='%H:%M:%S')
        df.HORA_DE_FINALIZACION=pd.to_datetime(df.HORA_DE_FINALIZACION, format='%H:%M:%S')
        ######usado si deseo generar insert mas facilmente
        
        # [ print('listAudits.{0} = None if pd.isnull(row.{0}) else row.{0}'.format(column)) for column in headers]
        
        
        df.ENVIAR=pd.to_numeric(df.ENVIAR,downcast="integer")
        df.ENVIAR=df.ENVIAR.astype('Int64')
        


        
        headers2=[]
        index=df.columns.array
        [headers2.append(column.strip().replace("_"," ")) for column in index]
        df.columns=headers2

        log.insertLogToDB('Success','prepareFieldsToLoad','Success execution in Mod Audit_Units')
        return df    
        

    except Exception as e:
        log.insertLogToDB(str(e),'prepareFieldsToLoad','Failed execution in Mod Audit_Units')
        raise

def populationFieldsOfBusiness(listAudits:list,date:datetime,fecha_carga):
    try:
        
        #fecha_carga=datetime.today()

        listAudits.ESTADO_GENERAL= 'SIN GESTION' if listAudits.ESTADO_AUDITORIA==None and listAudits.TIMESTAMP==None else ('AUDITADO' if listAudits.ESTADO_AUDITORIA =='AUDITABLE' else 'NO AUDITADO')
        listAudits.TIEMPO_EN_MIN =None if listAudits.HORA_FINALIZACION== None or listAudits.HORA_INICIO== None else (listAudits.HORA_FINALIZACION-listAudits.HORA_INICIO)*24*60      
        listAudits.FECHA_CARGA=fecha_carga
        listAudits.AÑO=date.year
        listAudits.MES=date.month
        listAudits.FECHA=date   
        listAudits.MESA_ASIG=listAudits.MESA

        return(listAudits)

    except Exception as e:
        log.insertLogToDB(str(e),'populationFieldsOfBusiness','Failed execution in Mod Audit_Units')
        raise


def LoadAuditsUnits(TmpTable,DimCampos):
    try:
        ###### Obtiene datos de la base de datos
        data=findByEstadoGeneral_SinGestion_And_errorDescripcion_NotLikeDup_or_null(TmpTable,DimCampos)

        #####Obtiene el path donde esta el excel y lo crea dinamicamente
        path=os.path.sep.join([pathOneDrive,"appsheet",appSheetFolderName, excelName])
        
        
        ######  Carga el workbook, elimina la hoja de excel, vuelve a crearla y mantiene toda la informacion de las otras hojas 
        book = load_workbook(path)
        sheet = book[pedidoExcelSheet] 
        book.remove_sheet(sheet)
        book.create_sheet(pedidoExcelSheet,0)
        # writer = pd.ExcelWriter(path, engine='openpyxl') 
        if not os.path.exists(pathOneDrive):
            os.makedirs(pathOneDrive)
        writer = pd.ExcelWriter(pathOneDrive+os.sep+excelName, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        ###### Ingresa toda la informacion en el excel
        data.to_excel(writer, pedidoExcelSheet,index=False)

        ###### Guarda el archivo
        writer.save()

        
        log.insertLogToDB('Success','LoadAuditsUnits','Success execution in Mod Audit_Units')

    except Exception as e:
        log.insertLogToDB(str(e),'LoadAuditsUnits','Failed execution in Mod Audit_Units')
        raise

def findByEstadoGeneral_SinGestion_And_errorDescripcion_NotLikeDup_or_null (TmpTable,DimCampos):
    try:
        
        with session.begin():
            campos=findAllByTIPO_CAMPO(('GENERAL','PRUEBA','INTERNO'),DimCampos)
            COLUMNNAMES=[]
            renamedColumns={}
            for campo in campos:
                COLUMNNAMES.append(campo.NOMBRE_EN_BD)
                renamedColumns[campo.NOMBRE_EN_BD]=campo.NOMBRE_EN_EXCEL

            params=session.query(TmpTable).filter(
                        and_(TmpTable.ESTADO_GENERAL == 'SIN GESTION',
                        or_(TmpTable.ERROR_DESCRIPCION.notlike('%Dup%'),
                        TmpTable.ERROR_DESCRIPCION == None)) ).statement
            
            ###### Crea el dataframe utilizando la query creada anteriormente y los nombres de las columnas a leer
            df = pd.DataFrame(pd.read_sql_query( sql=params , con=base.engine), columns =COLUMNNAMES)
            
            ##### renombra las columnas de la base de datos al formato necesario de el excel
        
            
            df2=df.rename(columns=renamedColumns)

            return df2
    except Exception as e:
        log.insertLogToDB(str(e),'findByEstadoGeneral_SinGestion_And_errorDescripcion_NotLikeDup_or_null','Failed execution in Mod Audit_Units')
        raise

def main(readingDate,strTmpTable,TmpTable,DimCampos):
    ##readExcelSrc es la funcion que lee la fuente de datos de excel
    result_df = readExcelSrc(readingDate)    
    ##truncateFactTable es la funcion que trunca la tabla temporal para poder cargar los datos nuevos a procesar
    truncateTmpTable(strTmpTable)    
    ##prepareFieldsToLoad prepara algunos campos de fecha y horas y los castea acorde a su tipo en la BD para que puedan ser cargados sin error
    result_df_prepared = prepareFieldsToLoad(result_df)
    ##inserta los datos a la tabla temporal y ademas popula algunos campos de negocio a partir de la evaluación de otros campos
    insertUAToTmpTable(result_df_prepared,readingDate,DimCampos,TmpTable) 
